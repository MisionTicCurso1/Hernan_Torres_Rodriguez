from datetime import datetime
from openpyxl import load_workbook
ruta = "/run/media/nanchixx/Win Ex/hetor/Documents/MISION TIC/Hernan_Torres_Rodriguez/Codigo/ejercicioCrud.xlsx"
ruta = r"/run/media/nanchixx/Win Ex/hetor/Documents/MISION TIC/Hernan_Torres_Rodriguez/Codigo/ejercicioCrud.xlsx"
def agregar(ruta, dato):
    archivo_excel = load_workbook(ruta)
    hojas_datos = archivo_excel["pacientes"]
    hojas_datos = hojas_datos["A2": "H" + str(hojas_datos.max_row + 1)]
    hojas = archivo_excel.active
    nombre = 2
    apellido = 3
    grupoSanguineo = 4
    Rh = 5
    edad = 6
    vacunado = 7
    contagiado = 8
    print(dato)
    for i in hojas_datos:
        print(i[0].value)
        if not(isinstance(i[0], int)):
            identificador = i[0].row
            hojas.cell(row= identificador, column= 1).value = identificador -1
            hojas.cell(row= identificador, column= nombre).value = dato["nombre"]
            hojas.cell(row= identificador, column= apellido).value = dato["apellido"]
            hojas.cell(row= identificador, column= grupoSanguineo).value = dato["grupoSanguineo"]
            hojas.cell(row= identificador, column= Rh).value = dato["Rh"] 
            hojas.cell(row= identificador, column= edad).value = dato["edad"]
            hojas.cell(row= identificador, column= vacunado).value = dato["vacunado"]
            hojas.cell(row= identificador, column= contagiado).value = dato["contagiado"]
            break
    archivo_excel.save(ruta)

    return
def leer(ruta, extraer):
    # print("d")
    Archivo_Excel = load_workbook(ruta)
    Hojas_datos = Archivo_Excel["pacientes"]
    Hojas_datos = Hojas_datos["A2": "H"+str(Hojas_datos.max_row)]
   
    info = {}

    for i in Hojas_datos:
        if ( isinstance(i[0].value, int)):
            info.setdefault(i[0].value, {'nombre': i[1].value, 'apellido': i[2].value, "grupoSanguineo": i[3].value,"Rh": i[4].value, "edad": i[5].value,"vacunado": i[6].value, "contagiado": i[7].value})
    if not(extraer == 'todo'):
        info = filtrar(info, extraer)
    for i in info:
        print()
        print()
        print("**** Paciente *****")
        print("Id: " + str(i) , '\n')
        print("Nombre:" + info[i]['nombre'] , '\n')
        print("Apellido:" + info[i]['apellido'] , '\n')
        print("grupo Sanguineo:" + info[i]['grupoSanguineo']  , '\n')
        print("RH: " + info[i]['Rh']  , '\n')
        print("Edad: " , info[i]['edad']  , '\n')
        print("Vacunado:" + info[i]['vacunado']  , '\n')
        print("Contagiado:" + info[i]['contagiado']  , '\n')
        print()
        print()

        
    return
def filtrar(info, filtro):
    temp = {}
    if(filtro == 'positivo' or filtro == 'negativo'):
        for i in info:
            if info[i]['Rh'] == filtro:
                temp.setdefault(i, info[i])
    elif(filtro == 'si' or filtro == 'no'):
        for i in info:
            if info[i]['vacunado'] == filtro:
                temp.setdefault(i, info[i])
    elif(filtro == 'Si' or filtro == 'No'):
        for i in info:
            if info[i]['contagiado'] == filtro:
                temp.setdefault(i, info[i])
    return temp

def borrar(ruta, identificador):
    Archivo_Excel = load_workbook(ruta)
    Hoja_Excel = Archivo_Excel['pacientes']
    Hoja_Excel = Hoja_Excel['A2': 'H'+str(Hoja_Excel.max_row)]
    hojas = Archivo_Excel.active

    nombre = 2
    apellido = 3
    grupoSanguineo = 4
    Rh = 5
    edad = 6
    vacunado = 7
    contagiado = 8

    for i in Hoja_Excel:
        if i[0].value == identificador:
            fila = i[0].row
            encontro = True
            hojas.cell(row=fila, column=1).value = ""
            hojas.cell(row=fila, column=nombre).value = ""
            hojas.cell(row=fila, column=apellido).value = ""
            hojas.cell(row=fila, column=grupoSanguineo).value = ""
            hojas.cell(row=fila, column=Rh).value = ""
            hojas.cell(row=fila, column=edad).value = ""
            hojas.cell(row=fila, column=vacunado).value = ""
            hojas.cell(row=fila, column=contagiado).value = ""
    Archivo_Excel.save(ruta)
    print(f'El paciente {identificador}, ha sido eliminado con exito')
    if(encontro == False):
        print("No existe ese ID")
        print()
    return
while(True):
    print('Indique la accion que quiere realizar')
    print()
    print('Consultar Paciente: 1')
    print('Actualizar Datos del paciente: 2')
    print('Agregar paciente: 3')
    print('Borrar Paciente: 4')
    print('Salir del sistema: 9')
    print()
    accion = int(input())
    if not (accion == 1 or accion == 2 or accion == 3 or accion == 4 or accion == 9):
        print('Ingresaste un digito incorrecto \n')
    if( accion == 1 ):
        print('--Ingresaste a la acción Consultar Paciente--')
        print('Filtrar por: \n RH: 1 \n Vacunados:2 \n Infectados:3 ')
        seleccion = int(input())
        if(seleccion == 1):
            print('--Filtar por RH--')
            print('Seleccione el RH a consultar\n Positivo: 1 \n Negativo: 2')
            rh = int(input())
            if(rh == 1):
                print('--Filtar por RH Positivo--')
                leer(ruta,'positivo')
            elif(rh == 2):
                print('--Filtar por RH Negativo--')
                leer(ruta,'negativo')
        elif(seleccion == 2):
            print('--Filtar por Vacunados--')
            print('Filtrar por los Si Vacunados : 1 \n No vacunados : 2')
            seleccion = int(input())
            if(seleccion == 1):
                leer(ruta,'si')
            elif(seleccion == 2):
                leer(ruta,'no')
        elif(seleccion == 3):
            print('--Filtar por Infectados--')
            leer(ruta,'Si')
    elif( accion == 2 ):
        print('--Ingresaste a la acción Actualizar Datos Paciente--')
    elif( accion == 3 ):
        print('--Ingresaste a la acción Agregar Paciente--')
        diccionario = {}
        diccionario.setdefault("nombre", input('Ingrese el nombre del paciente \n'))
        diccionario.setdefault("apellido", input('Ingrese el apellido del paciente \n'))
        print('Seleccione el Grupo Sanguineo del paciente \n A: 1 \n O: 2 \n AB: 3 \n')
        grupo = int(input())
        if(grupo == 1):
            diccionario.setdefault("grupoSanguineo",'A')
            print('Seleccione el RH del paciente \n positivo: 1 \n negativo: 2')
            RhS = int(input())
            if(RhS == 1):
                diccionario.setdefault("Rh",'positivo')
            elif(RhS == 2):
                diccionario.setdefault("Rh",'negativo')
        elif(grupo == 2):
            diccionario.setdefault("grupoSanguineo",'O')
            print('Seleccione el RH del paciente \n positivo: 1 \n negativo: 2')
            RhS = int(input())
            if(RhS == 1):
                diccionario.setdefault("Rh",'positivo')
            elif(RhS == 2):
                diccionario.setdefault("Rh",'negativo')
        elif(grupo == 3):
            diccionario.setdefault("grupoSanguineo",'AB')
            print('Seleccione el RH del paciente \n positivo: 1 \n negativo: 2')
            RhS = int(input())
            if(RhS == 1):
                diccionario.setdefault("Rh",'positivo')
            elif(RhS == 2):
                diccionario.setdefault("Rh",'negativo')
        diccionario.setdefault("edad", input('Ingrese la edad del paciente\n'))
        diccionario.setdefault("vacunado", (input('¿El paciente se encuentra vacunado? (si,no)\n').lower()))
        diccionario.setdefault("contagiado", (input('El paciente se ha contagiado por COVID19 (si,no)\n').lower()))
        agregar(ruta,diccionario)
    elif( accion == 4 ):
        print('--Ingresaste a la acción Borrar Paciente--')
        id = int(input('Ingrese el ID del paciente a eliminar \n'))
        borrar(ruta,id)
    elif( accion == 9 ):
        print('\n Gracias por usar el sistema... \n')
        break